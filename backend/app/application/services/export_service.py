import io
from datetime import datetime
from decimal import Decimal


class ExportService:
    @staticmethod
    def export_to_excel(data: list[dict], filename: str = "report") -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        if not data:
            return io.BytesIO().getvalue()

        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                value = row_data[key]
                if isinstance(value, Decimal):
                    value = float(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def export_to_pdf(data: list[dict], title: str = "Report", headers: list[str] = None) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"],
            fontSize=16, alignment=1, spaceAfter=20
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))

        if not data:
            elements.append(Paragraph("No data available", styles["Normal"]))
            doc.build(elements)
            return buffer.getvalue()

        if not headers:
            headers = list(data[0].keys())

        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        ]))

        elements.append(table)
        doc.build(elements)
        return buffer.getvalue()

    @staticmethod
    def export_to_csv(data: list[dict], filename: str = "report") -> str:
        import csv
        if not data:
            return ""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue()
